from pathlib import Path
import sqlite3
import pandas as pd
import numpy as np

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DB_PATH = PROJECT_ROOT / "db" / "nifty100.db"
OUTPUT_DIR = PROJECT_ROOT / "output"

METRIC_COLUMNS = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "free_cash_flow_cr",
    "pat_cagr_5yr",
    "revenue_cagr_5yr",
    "eps_cagr_5yr",
    "interest_coverage",
    "asset_turnover",
    "cfo_pat_ratio",
    "fcf_conversion_rate_pct",
    "earnings_per_share",
    "book_value_per_share",
    "dividend_payout_ratio_pct",
    "total_debt_cr",
    "revenue_cagr_3yr",
    "pat_cagr_3yr",
    "composite_quality_score",
    "cfo_quality_label",
]


def load_latest_ratios(conn):
    return pd.read_sql_query(
        """
        SELECT fr.*
        FROM financial_ratios fr
        INNER JOIN (
            SELECT company_id, MAX(year) AS latest_year
            FROM financial_ratios
            GROUP BY company_id
        ) latest
        ON fr.company_id = latest.company_id AND fr.year = latest.latest_year
        """,
        conn,
    )


def run():
    if not DB_PATH.exists():
        raise FileNotFoundError(f"Database not found: {DB_PATH}")

    conn = sqlite3.connect(DB_PATH)

    try:
        # companies table uses 'id' as the company identifier
        companies = pd.read_sql_query("SELECT id AS company_id, company_name FROM companies", conn)
        peer_groups = pd.read_sql_query("SELECT company_id, peer_group_name, is_benchmark FROM peer_groups", conn)
        ratios = load_latest_ratios(conn)
        percentiles = pd.read_sql_query("SELECT * FROM peer_percentiles", conn)

        merged = peer_groups.merge(companies, on="company_id", how="left").merge(ratios, on="company_id", how="left")

        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        out_path = OUTPUT_DIR / "peer_comparison.xlsx"

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:

            groups = sorted(merged["peer_group_name"].dropna().unique())

            for group in groups:
                group_df = merged[merged["peer_group_name"] == group].copy()

                # pick metric columns that exist
                cols = [c for c in METRIC_COLUMNS if c in group_df.columns]

                report_cols = ["company_id", "company_name"] + cols

                report = group_df[report_cols].copy()

                # attach percentiles for each metric
                for metric in cols:
                    pcol = metric + "_percentile"
                    # join from percentiles table
                    pvals = percentiles[percentiles["metric"] == metric][["company_id", "percentile_rank"]]
                    pvals = pvals.rename(columns={"percentile_rank": pcol})
                    report = report.merge(pvals, on="company_id", how="left")

                # summary median row
                median = report[cols].median(numeric_only=True)
                summary = {c: median.get(c, np.nan) for c in cols}
                summary_row = {"company_id": "MEDIAN", "company_name": "Peer Median"}
                summary_row.update(summary)

                report.to_excel(writer, sheet_name=str(group)[:31], index=False)

        # Post-process for color-coding and highlight benchmark
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        wb = load_workbook(out_path)
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        gold = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            headers = [cell.value for cell in ws[1]]

            # find percentile columns by header ending
            for i, h in enumerate(headers, start=1):
                if isinstance(h, str) and h.endswith("_percentile"):
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=i)
                        try:
                            v = float(cell.value) if cell.value is not None else None
                        except Exception:
                            v = None

                        if v is None:
                            continue
                        if v >= 75:
                            cell.fill = green
                        elif v > 25:
                            cell.fill = yellow
                        else:
                            cell.fill = red

            # highlight benchmark row if exists in peer_groups
            # attempt to find benchmark company from DB
            # load peer_groups for this sheet
            pg = pd.read_sql_query("SELECT company_id, peer_group_name, is_benchmark FROM peer_groups", conn)
            bench = pg[(pg['peer_group_name'] == sheet) & (pg['is_benchmark'] == 1)]['company_id']
            if not bench.empty:
                bench_id = bench.iloc[0]
                # find row with that company id
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=1).value == bench_id:
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = gold
                        break

        wb.save(out_path)

        print(f"Peer comparison written: {out_path}")

    finally:
        conn.close()


if __name__ == '__main__':
    run()
