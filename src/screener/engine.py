from pathlib import Path
import sqlite3
import yaml
import numpy as np
import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parents[2]

DB_PATH = PROJECT_ROOT / "db" / "nifty100.db"
OUTPUT_DIR = PROJECT_ROOT / "output"
CONFIG_PATH = PROJECT_ROOT / "config" / "screener_config.yaml"

# Map config metric keys to actual financial_ratios DB column names
METRIC_ALIAS = {
    "fcf": "free_cash_flow_cr",
    "fcf_latest_positive": "free_cash_flow_cr",
    "cfo_pat_ratio": "cfo_pat_ratio",
    "revenue_cagr_5yr": "revenue_cagr_5yr",
    "revenue_cagr_3yr": "revenue_cagr_3yr",
    "pat_cagr_5yr": "pat_cagr_5yr",
    "eps_cagr_5yr": "eps_cagr_5yr",
    "interest_coverage": "interest_coverage",
    "debt_to_equity": "debt_to_equity",
    "dividend_payout_pct": "dividend_payout_ratio_pct",
    "dividend_yield_pct": "dividend_yield_pct",
    "revenue": "revenue",
}


def load_config(path=CONFIG_PATH):
    with open(path, "r", encoding="utf8") as fh:
        return yaml.safe_load(fh)


def load_latest_financial_ratios(conn):
    df = pd.read_sql_query(
        """
        SELECT fr.*
        FROM financial_ratios fr

        INNER JOIN
        (
            SELECT company_id, MAX(year) AS latest_year
            FROM financial_ratios
            GROUP BY company_id
        ) latest
        ON fr.company_id = latest.company_id
           AND fr.year = latest.latest_year
        """,
        conn,
    )
    return df


def load_sectors(conn):
    return pd.read_sql_query("SELECT company_id, broad_sector FROM sectors", conn)


def compute_flags_from_history(conn):
    """
    Compute helper boolean flags: fcf_latest_positive and debt_to_equity_declining
    Returns dicts keyed by company_id.
    """
    # use actual column name for free cash flow
    history = pd.read_sql_query(
        "SELECT company_id, year, debt_to_equity, free_cash_flow_cr FROM financial_ratios ORDER BY company_id, year DESC",
        conn,
    )

    flags = {}

    for comp, group in history.groupby("company_id"):
        group = group.sort_values("year", ascending=False)
        latest = group.iloc[0]
        fcf_pos = pd.to_numeric(latest.get("fcf"), errors="coerce") > 0
        debt_decline = False

        if len(group) >= 2:
            prev = group.iloc[1]
            try:
                debt_decline = (
                    float(latest.get("debt_to_equity") or np.nan)
                    < float(prev.get("debt_to_equity") or np.nan)
                )
            except Exception:
                debt_decline = False

        flags[comp] = {
            "fcf_latest_positive": bool(fcf_pos),
            "debt_to_equity_declining": bool(debt_decline),
        }

    return flags


def _prepare_numeric(series, treat_debt_free_icr=False):
    if treat_debt_free_icr:
        # Treat string 'Debt Free' as infinity
        series = series.replace("Debt Free", np.inf)

    return pd.to_numeric(series, errors="coerce")


def apply_filters(df, sectors, thresholds, history_flags=None):
    """
    Apply threshold filters to the provided financial ratios DataFrame.
    Returns filtered DataFrame and boolean mask columns for each filter.
    """
    df = df.copy()
    sectors = sectors.copy()
    df["company_id"] = df["company_id"].astype(str)
    sectors["company_id"] = sectors["company_id"].astype(str)

    # Merge sectors for D/E skipping logic
    merged = df.merge(sectors, on="company_id", how="left")

    mask = pd.Series(True, index=merged.index)
    pass_fail = {}

    for metric, rule in thresholds.items():
        col = metric

        # handle special flags
        if metric in {"fcf_latest_positive", "debt_to_equity_declining"}:
            # historical flags are stored per company
            values = merged["company_id"].map(lambda cid: history_flags.get(cid, {}).get(metric, False) if history_flags else False)
            meets = values == rule.get("equals", True)
            pass_fail[metric] = meets
            mask &= meets
            continue

        # map metric name to DB column
        db_col = METRIC_ALIAS.get(metric, metric)

        # prepare numeric series
        if metric == "interest_coverage":
            series = _prepare_numeric(merged.get(db_col), treat_debt_free_icr=True)
        else:
            series = _prepare_numeric(merged.get(db_col))

        # ensure series is a pandas Series even if column missing
        if not isinstance(series, pd.Series):
            series = pd.Series([series] * len(merged), index=merged.index)

        # D/E filter skip for Financials broad sector
        if metric == "debt_to_equity" and "max" in rule:
            max_val = float(rule["max"])
            # companies in Financials automatically pass
            broad = merged.get("broad_sector")
            if broad is None:
                is_financial = pd.Series(False, index=merged.index)
            else:
                is_financial = broad == "Financials"

            meets = (series <= max_val) | is_financial.fillna(False)
            pass_fail[metric] = meets.fillna(False)
            mask &= meets.fillna(False)
            continue

        if "min" in rule:
            meets = series >= float(rule["min"])
        elif "max" in rule:
            meets = series <= float(rule["max"])
        elif "eq" in rule:
            meets = series == float(rule["eq"])
        else:
            # fallback exact equals for booleans
            if "equals" in rule:
                meets = series == rule["equals"]
            else:
                meets = pd.Series(False, index=series.index)

        pass_fail[metric] = meets.fillna(False)
        mask &= meets.fillna(False)

    filtered = merged[mask]

    # add columns matching config metric names (so exports and colouring can refer to config keys)
    for metric in thresholds.keys():
        db_col = METRIC_ALIAS.get(metric, metric)
        if db_col in merged.columns:
            filtered[metric] = filtered[db_col]
        else:
            filtered[metric] = np.nan

    return filtered, pass_fail


def winsorise_and_scale(series, lower_pct=10, upper_pct=90, invert=False):
    numeric = pd.to_numeric(series, errors="coerce")
    if numeric.dropna().empty:
        return pd.Series(np.nan, index=series.index)

    lower = np.nanpercentile(numeric, lower_pct)
    upper = np.nanpercentile(numeric, upper_pct)
    capped = numeric.clip(lower, upper)
    # scale to 0-100
    minv = capped.min()
    maxv = capped.max()
    if pd.isna(minv) or pd.isna(maxv) or maxv == minv:
        scaled = pd.Series(50.0, index=series.index)
    else:
        scaled = (capped - minv) / (maxv - minv) * 100.0

    if invert:
        scaled = 100.0 - scaled

    return scaled


def compute_composite_score(df):
    """
    Compute composite quality score per specification and add `composite_quality_score` column.
    """
    work = df.copy()

    # Define metric mapping and weights using DB column names
    roe = winsorise_and_scale(work["return_on_equity_pct"] if "return_on_equity_pct" in work.columns else pd.Series(np.nan, index=work.index))
    roce = winsorise_and_scale(work["return_on_capital_employed_pct"] if "return_on_capital_employed_pct" in work.columns else pd.Series(np.nan, index=work.index))
    npm = winsorise_and_scale(work["net_profit_margin_pct"] if "net_profit_margin_pct" in work.columns else pd.Series(np.nan, index=work.index))

    fcf_series = work[METRIC_ALIAS.get("fcf")] if METRIC_ALIAS.get("fcf") in work.columns else pd.Series(np.nan, index=work.index)
    fcf = winsorise_and_scale(fcf_series)
    cfo_pat = winsorise_and_scale(work["cfo_pat_ratio"] if "cfo_pat_ratio" in work.columns else pd.Series(np.nan, index=work.index))
    fcf_flag = (fcf_series > 0).astype(float)
    fcf_flag_score = fcf_flag * 100.0 * 0.05  # 5% of total

    revenue_cagr = winsorise_and_scale(work["revenue_cagr_5yr"] if "revenue_cagr_5yr" in work.columns else pd.Series(np.nan, index=work.index))
    pat_cagr = winsorise_and_scale(work["pat_cagr_5yr"] if "pat_cagr_5yr" in work.columns else pd.Series(np.nan, index=work.index))

    de_score = winsorise_and_scale(work["debt_to_equity"] if "debt_to_equity" in work.columns else pd.Series(np.nan, index=work.index), invert=True)
    icr_score = winsorise_and_scale(work["interest_coverage"] if "interest_coverage" in work.columns else pd.Series(np.nan, index=work.index))

    # Compose weighted sub-scores
    profitability = (roe * 0.15 + roce * 0.10 + npm * 0.10) * (0.35 / 0.35)
    cash_quality = (fcf * 0.15 + cfo_pat * 0.10 + fcf_flag_score) * (0.30 / 0.30)
    growth = (revenue_cagr * 0.10 + pat_cagr * 0.10) * (0.20 / 0.20)
    leverage = (de_score * 0.10 + icr_score * 0.05) * (0.15 / 0.15)

    composite = (
        (profitability.fillna(0))
        + (cash_quality.fillna(0))
        + (growth.fillna(0))
        + (leverage.fillna(0))
    )

    # Some rows may be NaN entirely; set to NaN
    composite[(roe.isna()) & (roce.isna()) & (npm.isna()) & (fcf.isna())] = np.nan

    work["composite_quality_score_raw"] = composite

    # Sector-relative normalization: scale within broad_sector to 0-100
    raw = work["composite_quality_score_raw"]
    if raw.dropna().empty:
        work["composite_quality_score"] = np.nan
    else:
        if "broad_sector" in work.columns:
            def scale_group(s):
                vals = s.dropna()
                if vals.empty:
                    return pd.Series(np.nan, index=s.index)
                lo = np.nanmin(vals)
                hi = np.nanmax(vals)
                if lo == hi:
                    return pd.Series(50.0, index=s.index)
                return pd.Series(np.interp(s.fillna((lo+hi)/2), (lo, hi), (0, 100)), index=s.index)

            work["composite_quality_score"] = work.groupby("broad_sector")["composite_quality_score_raw"].transform(scale_group)

        else:
            vals = raw.dropna()
            if vals.empty:
                work["composite_quality_score"] = np.nan
            else:
                lo = np.nanmin(vals)
                hi = np.nanmax(vals)
                if lo == hi:
                    work["composite_quality_score"] = 50.0
                else:
                    work["composite_quality_score"] = np.interp(raw.fillna((lo+hi)/2), (lo, hi), (0, 100))

    return work


def export_screener_outputs(all_results, config):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / "screener_output.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for preset_name, df in all_results.items():
            df.to_excel(writer, sheet_name=preset_name[:31], index=False)

    # Post-process for color coding
    wb = load_workbook(out_path)

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for preset_name, preset in config.get("presets", {}).items():
        sheet_name = preset_name[:31]
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        # For each filter metric, color cells
        for metric in preset.get("filters", {}).keys():
            if metric in headers:
                col_idx = headers.index(metric) + 1
                # iterate rows
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    try:
                        val = cell.value
                        meets = True
                        rule = preset["filters"][metric]
                        if "min" in rule:
                            meets = (val is not None) and (float(val) >= float(rule["min"]))
                        elif "max" in rule:
                            meets = (val is not None) and (float(val) <= float(rule["max"]))
                        elif "eq" in rule:
                            meets = (val is not None) and (float(val) == float(rule["eq"]))
                        else:
                            meets = True

                    except Exception:
                        meets = False

                    cell.fill = green if meets else red

    wb.save(out_path)


def run_all():
    config = load_config()

    if not DB_PATH.exists():
        raise FileNotFoundError(f"Database not found: {DB_PATH}")

    conn = sqlite3.connect(DB_PATH)

    try:
        ratios = load_latest_financial_ratios(conn)
        sectors = load_sectors(conn)
        history_flags = compute_flags_from_history(conn)

        # compute composite score on full universe first
        merged = ratios.merge(sectors, on="company_id", how="left")
        merged = compute_composite_score(merged)

        results = {}

        for preset_name, preset in config.get("presets", {}).items():
            thresholds = preset.get("filters", {})
            filtered, pass_fail = apply_filters(merged, sectors, thresholds, history_flags)
            # ensure composite present
            filtered = compute_composite_score(filtered)
            # Keep KPI columns and composite
            results[preset_name] = filtered

        export_screener_outputs(results, config)

    finally:
        conn.close()


if __name__ == "__main__":
    run_all()
